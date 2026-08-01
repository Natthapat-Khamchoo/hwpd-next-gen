"""
ออกรายงานตามแบบฟอร์มใน `tb_ReportCatalog` เป็นไฟล์ Excel

รองรับสองแบบ:

  summary  ตารางยอดรวมรายกองกำกับ ใช้ตัวเลขจาก tb_National_Summary
  tags     ตารางยอดตามป้ายหมวด (ECIG, GUN, DRUG, WEIGHT ฯลฯ) อ่านจาก tb_ReportCache
  cases    รายการคดีทีละใบ กรองด้วยป้ายหมวดของข้อหา สแกน tb_Arrests สด

รายการแบบฟอร์มอ่านจาก tb_ReportCatalog ไม่ได้ฮาร์ดโค้ด เพิ่มในชีตแล้วโผล่เอง
เหลือชนิด บรรยาย ที่ยังไม่รองรับ เพราะเป็นเอกสารบรรยาย/PPT ไม่ใช่ตาราง

ดูสถานะรายแบบฟอร์มได้จาก available_reports()
"""

import io
import logging
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.config import MASTER_SHEET_ID, get_db_router, get_target_db_id
from app.services import (
    national_service,
    query_service,
    report_cache_service,
    sheets_service,
    tag_stats_service,
)

logger = logging.getLogger(__name__)


class ReportNotSupported(RuntimeError):
    """แบบฟอร์มนี้ยังออกไม่ได้"""


# reportKey -> รายละเอียดการออกรายงาน
# ต้องตรงกับ reportKey ใน tb_ReportCatalog
SUPPORTED = {
    "RPT_FRIDAY_STATS": {
        "title": "สถิติผลการปฏิบัติ (ส่งห้องขับเคลื่อน)",
        "cadence": "ทุกวันศุกร์",
    },
}

# ชนิดที่ออกเป็นตารางยอดรวมตามป้ายหมวด
TAG_REPORT_TYPES = {"ยอดรวม", "รายกิจกรรม", "ทะเบียนบุคคล"}

# ชนิดที่ออกเป็นรายการคดีทีละใบ ไม่ใช่ยอดรวม
CASE_REPORT_TYPES = {"รายคดี"}

# ของเดิมจำกัดไว้เท่านี้ กันไฟล์ใหญ่เกินและกันสแกนทั้งชีตนานเกินไป
CASE_LIMIT = 500

# คอลัมน์ในไฟล์รายคดี -> ชื่อคอลัมน์ใน tb_Arrests
# ชุดเดียวกับที่ getReportCases ของเดิมส่งกลับ
CASE_COLUMNS = [
    ("วันที่", query_service.COL_ACTUAL_DATE),
    ("กก.", None),
    ("สถานี", query_service.COL_STATION_ID),
    ("หน่วย", query_service.COL_UNIT_ID),
    ("เลขคดี", "เลขคดี"),
    ("หัวข้อการจับกุม", "หัวข้อการจับกุม"),
    ("จับโดย", "จับโดย"),
    ("ประเภทการจับกุม", "ประเภทการจับกุม"),
    ("ขอบเขตหมาย", "ขอบเขตหมาย"),
    ("ตรวจค้น/แจ้งข้อกล่าวหา", "ตรวจค้น/แจ้งข้อกล่าวหา"),
    ("วันเวลาที่ดำเนินการ", "วันที่เวลาที่ดำเนินการ"),
    ("ชุดจับกุม", "ชุดจับกุม - เก็บเป็นรายชื่อต่อกันด้วยลูกน้ำ"),
    ("จำนวนผู้ต้องหา", "จำนวนผู้ต้องหา"),
    ("ผู้ต้องหา", "ข้อมูลผู้ต้องหาทั้งหมด"),
    ("ข้อหา", "ข้อหาทั้งหมด"),
    ("สถานที่จับกุม", "สถานที่จับกุม"),
    ("ของกลาง", "ของกลาง"),
    ("ประเภทคดีบุหรี่ไฟฟ้า", "ประเภทคดีบุหรี่ไฟฟ้า"),
    ("เว็บไซต์/URL", "เว็บไซต์/URL ที่เกี่ยวข้อง"),
    ("มูลค่าความเสียหาย", "มูลค่าความเสียหาย (บาท)"),
    ("วงเงินหมุนเวียน", "วงเงินหมุนเวียน (บาท)"),
    ("พฤติการณ์", "พฤติการณ์"),
    ("การดำเนินการส่งต่อ", "การดำเนินการส่งต่อ"),
    ("recordId", query_service.COL_RECORD_ID),
]

# ชื่อฟิลด์ใน national_summary -> หัวคอลัมน์บนรายงาน
METRICS = [
    ("arrestsCount", "จับกุมคดีอาญา"),
    ("v20Count", "คดีจราจร (ว.20)"),
    ("v43Count", "ว.43"),
    ("v42Count", "ว.42"),
    ("serviceCount", "บริการ"),
    ("volCount", "จิตอาสา"),
    ("royalCount", "รับเสด็จ"),
    ("missionCount", "ภารกิจ"),
    ("accCount", "อุบัติเหตุ"),
    ("deadCount", "เสียชีวิต"),
    ("injuredCount", "บาดเจ็บ"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _catalog() -> List[Dict[str, str]]:
    rows = sheets_service.read_table(MASTER_SHEET_ID, "tb_ReportCatalog")
    if not rows:
        return []
    header = [str(h).strip() for h in rows[0]]
    out = []
    for row in rows[1:]:
        record = {c: (str(row[i]).strip() if i < len(row) else "") for i, c in enumerate(header)}
        if record.get("reportKey"):
            out.append(record)
    return out


def _tags_of(record: Dict[str, str]) -> List[str]:
    raw = record.get("dataTags (หมวดข้อมูล)", "")
    return [t.strip().upper() for t in raw.split(",") if t.strip()]


def available_reports() -> List[Dict[str, Any]]:
    """
    แบบฟอร์มที่ออกได้ตอนนี้

    อ่านจาก tb_ReportCatalog ไม่ได้ฮาร์ดโค้ดไว้ เพิ่มแบบฟอร์มในชีตแล้วโผล่เองถ้าชนิด
    รองรับและระบุ dataTags ไว้ ตัวที่ไม่มี dataTags ออกไม่ได้เพราะไม่รู้ว่าจะนับอะไร
    """
    reports = [
        {"reportKey": key, "title": spec["title"], "cadence": spec["cadence"],
         "kind": "summary", "tags": []}
        for key, spec in SUPPORTED.items()
    ]
    try:
        catalog = _catalog()
    except Exception as exc:
        logger.warning("อ่าน tb_ReportCatalog ไม่ได้: %s", exc)
        return reports

    for record in catalog:
        key = record["reportKey"]
        if key in SUPPORTED:
            continue
        if str(record.get("isActive", "")).strip().upper() == "FALSE":
            continue
        tags = _tags_of(record)
        kind_of = record.get("type (ชนิด)", "")
        if not tags:
            continue
        if kind_of in TAG_REPORT_TYPES:
            kind = "tags"
        elif kind_of in CASE_REPORT_TYPES:
            kind = "cases"
        else:
            continue
        reports.append({
            "reportKey": key,
            "title": record.get("reportName", key),
            "cadence": record.get("cadence (รอบส่ง)", ""),
            "kind": kind,
            "tags": tags,
        })
    return reports


def _header_row(worksheet, row: int, labels: List[str], first_width: int = 18) -> None:
    for index, label in enumerate(labels, start=1):
        cell = worksheet.cell(row=row, column=index, value=label)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    worksheet.row_dimensions[row].height = 32
    worksheet.column_dimensions["A"].width = first_width
    for index in range(2, len(labels) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 15


def _build_tag_report(report: Dict[str, Any], start: str, end: str) -> bytes:
    """ตารางยอดรายป้ายหมวด — แถวคือ กก. คอลัมน์คือ ป้าย x (ราย/คน/ใบสั่ง)"""
    tags = report["tags"]

    # อ่านจากแคชก่อน สแกนสดต้องเปิด 2 ตาราง x 8 กก. ซึ่งชนโควตา Google ได้ถ้ามีคน
    # กดพร้อมกัน ตกไปสแกนสดเฉพาะตอนแคชยังไม่ครอบคลุมวันไหนเลย
    cached = report_cache_service.read_range(start, end)
    if cached["cachedDates"]:
        by_division = cached["byDivision"]
        totals = cached["total"]
        missing = cached["missingDates"]
    else:
        data = tag_stats_service.aggregate_national_tags(start, end)
        by_division = data["byDivision"]
        totals = data["total"]
        missing = []

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ยอดตามหมวด"

    worksheet["A1"] = report["title"]
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = (f"ช่วงวันที่ {start} ถึง {end}   (รอบส่ง: {report['cadence']})   "
                       f"หมวดข้อมูล: {', '.join(tags)}")
    worksheet["A2"].font = Font(size=10, color="595959")

    # คดีอาญานับเป็นราย/คน ส่วนคดีจราจรนับเป็นใบสั่ง แยกคอลัมน์ให้ชัดว่าอันไหนคืออะไร
    labels = ["กองกำกับการ"]
    for tag in tags:
        labels += [f"{tag}\nราย", f"{tag}\nคน", f"{tag}\nใบสั่ง"]
    header_row = 4
    _header_row(worksheet, header_row, labels)

    line = header_row
    for division in sorted(by_division):
        line += 1
        worksheet.cell(row=line, column=1, value=f"กก.{division}").border = BORDER
        column = 2
        for tag in tags:
            values = by_division[division].get(tag, {})
            for field in ("cases", "suspects", "citations"):
                cell = worksheet.cell(row=line, column=column, value=int(values.get(field, 0) or 0))
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center")
                column += 1

    line += 1
    cell = worksheet.cell(row=line, column=1, value="รวมทั้งประเทศ")
    cell.font, cell.fill, cell.border = Font(bold=True), TOTAL_FILL, BORDER
    column = 2
    for tag in tags:
        values = totals.get(tag, {})
        for field in ("cases", "suspects", "citations"):
            cell = worksheet.cell(row=line, column=column, value=int(values.get(field, 0) or 0))
            cell.font, cell.fill, cell.border = Font(bold=True), TOTAL_FILL, BORDER
            cell.alignment = Alignment(horizontal="center")
            column += 1

    worksheet.cell(row=line + 2, column=1,
                   value="ราย/คน มาจากรายงานจับกุม ส่วนใบสั่งมาจากยอดข้อหาในผลการปฏิบัติประจำวัน")
    # บอกให้เห็นบนหน้ารายงานเลย ไม่งั้นยอดที่ขาดวันจะถูกอ่านว่าไม่มีผลงาน
    if missing:
        shown = ", ".join(missing[:8]) + (" ..." if len(missing) > 8 else "")
        cell = worksheet.cell(row=line + 3, column=1,
                              value=f"ยังไม่มีข้อมูลในแคช {len(missing)} วัน: {shown} — สั่งอัปเดตแคชแล้วออกใหม่")
        cell.font = Font(color="C00000", bold=True)
    worksheet.freeze_panes = f"B{header_row + 1}"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_case_report(report: Dict[str, Any], start: str, end: str) -> bytes:
    """
    รายการคดีทีละใบ กรองด้วยป้ายหมวดของข้อหา — เทียบเท่า getReportCases ของเดิม

    สแกน tb_Arrests สดทุก กก. ไม่ผ่านแคช เพราะแคชเก็บแต่ยอดรวมรายป้าย ไม่มีตัวคดี
    ของเดิมก็สแกนสดด้วยเหตุผลเดียวกัน
    """
    want = set(report["tags"])
    mapping = tag_stats_service.charge_tags()

    rows: List[Dict[str, Any]] = []
    truncated = False
    for division in sorted(d for d, entry in get_db_router().items() if d != "0" and entry.get("OPS")):
        if len(rows) >= CASE_LIMIT:
            truncated = True
            break
        try:
            records = query_service.read_rows(get_target_db_id(f"{division}0"), "tb_Arrests")
        except Exception as exc:
            logger.warning("อ่าน tb_Arrests ของ กก.%s ไม่ได้ ข้ามไป: %s", division, exc)
            continue

        for record in records:
            if len(rows) >= CASE_LIMIT:
                truncated = True
                break
            if str(record.get(query_service.COL_STATUS, "")) not in tag_stats_service.COUNTED_STATUSES:
                continue
            if not query_service.is_active(record.get(query_service.COL_IS_ACTIVE)):
                continue
            date_value = str(record.get(query_service.COL_ACTUAL_DATE, "") or "")[:10]
            if not date_value or (start and date_value < start) or (end and date_value > end):
                continue
            found = tag_stats_service._tags_in(record.get("ข้อหาทั้งหมด", ""), mapping)
            if want and not (found & want):
                continue
            record["_division"] = f"กก.{division}"
            rows.append(record)

    rows.sort(key=lambda r: str(r.get(query_service.COL_ACTUAL_DATE, "")), reverse=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "รายคดี"

    worksheet["A1"] = report["title"]
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = (f"ช่วงวันที่ {start} ถึง {end}   (รอบส่ง: {report['cadence']})   "
                       f"หมวดข้อมูล: {', '.join(report['tags'])}   พบ {len(rows)} คดี")
    worksheet["A2"].font = Font(size=10, color="595959")

    header_row = 4
    labels = [label for label, _ in CASE_COLUMNS]
    _header_row(worksheet, header_row, labels, first_width=12)
    for index in range(2, len(labels) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 22

    line = header_row
    for record in rows:
        line += 1
        for index, (_, column) in enumerate(CASE_COLUMNS, start=1):
            value = record["_division"] if column is None else record.get(column, "")
            cell = worksheet.cell(row=line, column=index, value=str(value or ""))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if not rows:
        worksheet.cell(row=header_row + 2, column=1,
                       value="ไม่พบคดีที่ตรงกับหมวดนี้ในช่วงวันที่ที่เลือก")
    if truncated:
        cell = worksheet.cell(row=line + 2, column=1,
                              value=f"แสดงสูงสุด {CASE_LIMIT} คดี — แคบช่วงวันที่ลงเพื่อดูให้ครบ")
        cell.font = Font(color="C00000", bold=True)

    worksheet.freeze_panes = f"A{header_row + 1}"
    stream = io.BytesIO()
    workbook.save(stream)
    logger.info("ออกรายงานรายคดี %s ช่วง %s ถึง %s: %d คดี", report["reportKey"], start, end, len(rows))
    return stream.getvalue()


def build_workbook(report_key: str, start: str, end: str) -> bytes:
    """คืนไฟล์ Excel เป็น bytes ให้ endpoint ส่งต่อ"""
    report = next((r for r in available_reports() if r["reportKey"] == report_key), None)
    if not report:
        raise ReportNotSupported(
            f"แบบฟอร์ม {report_key} ยังออกอัตโนมัติไม่ได้ — ต้องเป็นชนิดที่รองรับ "
            "และระบุ dataTags ไว้ใน tb_ReportCatalog"
        )
    if report["kind"] == "cases":
        return _build_case_report(report, start, end)
    if report["kind"] == "tags":
        logger.info("ออกรายงาน %s (ตามป้าย %s) ช่วง %s ถึง %s", report_key, report["tags"], start, end)
        return _build_tag_report(report, start, end)

    spec = SUPPORTED[report_key]
    data = national_service.national_summary(start, end)
    totals = data.get("totals", {})
    by_division = sorted(data.get("byDivision", []), key=lambda r: str(r.get("div", "")))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ผลการปฏิบัติ"

    worksheet["A1"] = spec["title"]
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = f"ช่วงวันที่ {start} ถึง {end}   (รอบส่ง: {spec['cadence']})"
    worksheet["A2"].font = Font(size=10, color="595959")

    header_row = 4
    headers = ["กองกำกับการ"] + [label for _, label in METRICS]
    for index, label in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=index, value=label)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    worksheet.row_dimensions[header_row].height = 32
    worksheet.column_dimensions["A"].width = 18
    for index in range(2, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 14

    line = header_row
    for record in by_division:
        line += 1
        worksheet.cell(row=line, column=1, value=record.get("divName", "")).border = BORDER
        for offset, (field, _) in enumerate(METRICS, start=2):
            cell = worksheet.cell(row=line, column=offset, value=int(record.get(field, 0) or 0))
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    line += 1
    cell = worksheet.cell(row=line, column=1, value="รวมทั้งประเทศ")
    cell.font, cell.fill, cell.border = Font(bold=True), TOTAL_FILL, BORDER
    for offset, (field, _) in enumerate(METRICS, start=2):
        cell = worksheet.cell(row=line, column=offset, value=int(totals.get(field, 0) or 0))
        cell.font, cell.fill, cell.border = Font(bold=True), TOTAL_FILL, BORDER
        cell.alignment = Alignment(horizontal="center")

    # ไม่มีแถวของ กก. ไหนเลยแปลว่ายังไม่มีใครรวมยอดในช่วงนั้น ไม่ใช่ว่ายอดเป็นศูนย์
    if not by_division:
        worksheet.cell(row=line + 2, column=1,
                       value="ไม่พบข้อมูลในช่วงวันที่นี้ — ตรวจว่าตัวตั้งเวลารวมยอดทำงานแล้วหรือยัง")

    worksheet.freeze_panes = f"A{header_row + 1}"

    stream = io.BytesIO()
    workbook.save(stream)
    logger.info("ออกรายงาน %s ช่วง %s ถึง %s (%d กก.)", report_key, start, end, len(by_division))
    return stream.getvalue()
