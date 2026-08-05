# -*- coding: utf-8 -*-
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding="utf-8")

STATIONS_DATA = {
    1: [
        {"id": "10", "sheet": "ฝอ.กก.1", "fullName": "ฝอ.กก.1 บก.ทล."},
        {"id": "11", "sheet": "ส.ทล.1 (อยุธยา)", "fullName": "ส.ทล.1 กก.1 บก.ทล. (อยุธยา)"},
        {"id": "12", "sheet": "ส.ทล.2 (สระบุรี)", "fullName": "ส.ทล.2 กก.1 บก.ทล. (สระบุรี)"},
        {"id": "13", "sheet": "ส.ทล.3 (ลพบุรี)", "fullName": "ส.ทล.3 กก.1 บก.ทล. (ลพบุรี)"},
        {"id": "14", "sheet": "ส.ทล.4 (นครสวรรค์)", "fullName": "ส.ทล.4 กก.1 บก.ทล. (นครสวรรค์)"},
        {"id": "15", "sheet": "ส.ทล.5 (เพชรบูรณ์)", "fullName": "ส.ทล.5 กก.1 บก.ทล. (เพชรบูรณ์)"},
        {"id": "16", "sheet": "ส.ทล.6 (สิงห์บุรี)", "fullName": "ส.ทล.6 กก.1 บก.ทล. (สิงห์บุรี)"}
    ],
    2: [
        {"id": "20", "sheet": "ฝอ.กก.2", "fullName": "ฝอ.กก.2 บก.ทล."},
        {"id": "21", "sheet": "ส.ทล.1 (นครปฐม)", "fullName": "ส.ทล.1 กก.2 บก.ทล. (นครปฐม)"},
        {"id": "22", "sheet": "ส.ทล.2 (เพชรบุรี)", "fullName": "ส.ทล.2 กก.2 บก.ทล. (เพชรบุรี)"},
        {"id": "23", "sheet": "ส.ทล.3 (ประจวบฯ)", "fullName": "ส.ทล.3 กก.2 บก.ทล. (ประจวบคีรีขันธ์)"},
        {"id": "24", "sheet": "ส.ทล.4 (ชุมพร)", "fullName": "ส.ทล.4 กก.2 บก.ทล. (ชุมพร)"},
        {"id": "25", "sheet": "ส.ทล.5 (สุราษฎร์ฯ)", "fullName": "ส.ทล.5 กก.2 บก.ทล. (สุราษฎร์ธานี)"},
        {"id": "26", "sheet": "ส.ทล.6 (สุพรรณบุรี)", "fullName": "ส.ทล.6 กก.2 บก.ทล. (สุพรรณบุรี)"}
    ],
    3: [
        {"id": "30", "sheet": "ฝอ.กก.3", "fullName": "ฝอ.กก.3 บก.ทล."},
        {"id": "31", "sheet": "ส.ทล.1 (ฉะเชิงเทรา)", "fullName": "ส.ทล.1 กก.3 บก.ทล. (ฉะเชิงเทรา)"},
        {"id": "32", "sheet": "ส.ทล.2 (ชลบุรี)", "fullName": "ส.ทล.2 กก.3 บก.ทล. (ชลบุรี)"},
        {"id": "33", "sheet": "ส.ทล.3 (ระยอง)", "fullName": "ส.ทล.3 กก.3 บก.ทล. (ระยอง)"},
        {"id": "34", "sheet": "ส.ทล.4 (จันทบุรี)", "fullName": "ส.ทล.4 กก.3 บก.ทล. (จันทบุรี)"},
        {"id": "35", "sheet": "ส.ทล.5 (ปราจีนบุรี)", "fullName": "ส.ทล.5 กก.3 บก.ทล. (ปราจีนบุรี)"}
    ],
    4: [
        {"id": "40", "sheet": "ฝอ.กก.4", "fullName": "ฝอ.กก.4 บก.ทล."},
        {"id": "41", "sheet": "ส.ทล.1 (ร้อยเอ็ด)", "fullName": "ส.ทล.1 กก.4 บก.ทล. (ร้อยเอ็ด)"},
        {"id": "42", "sheet": "ส.ทล.2 (ขอนแก่น)", "fullName": "ส.ทล.2 กก.4 บก.ทล. (ขอนแก่น)"},
        {"id": "43", "sheet": "ส.ทล.3 (อุดรธานี)", "fullName": "ส.ทล.3 กก.4 บก.ทล. (อุดรธานี)"},
        {"id": "44", "sheet": "ส.ทล.4 (เลย)", "fullName": "ส.ทล.4 กก.4 บก.ทล. (เลย)"},
        {"id": "45", "sheet": "ส.ทล.5 (สกลนคร)", "fullName": "ส.ทล.5 กก.4 บก.ทล. (สกลนคร)"}
    ],
    5: [
        {"id": "50", "sheet": "ฝอ.กก.5", "fullName": "ฝอ.กก.5 บก.ทล."},
        {"id": "51", "sheet": "ส.ทล.1 (ตาก)", "fullName": "ส.ทล.1 กก.5 บก.ทล. (ตาก)"},
        {"id": "52", "sheet": "ส.ทล.2 (ลำปาง)", "fullName": "ส.ทล.2 กก.5 บก.ทล. (ลำปาง)"},
        {"id": "53", "sheet": "ส.ทล.3 (พิษณุโลก)", "fullName": "ส.ทล.3 กก.5 บก.ทล. (พิษณุโลก)"},
        {"id": "54", "sheet": "ส.ทล.4 (เชียงใหม่)", "fullName": "ส.ทล.4 กก.5 บก.ทล. (เชียงใหม่)"},
        {"id": "55", "sheet": "ส.ทล.5 (พะเยา)", "fullName": "ส.ทล.5 กก.5 บก.ทล. (พะเยา)"},
        {"id": "56", "sheet": "ส.ทล.6 (แพร่)", "fullName": "ส.ทล.6 กก.5 บก.ทล. (แพร่)"}
    ],
    6: [
        {"id": "60", "sheet": "ฝอ.กก.6", "fullName": "ฝอ.กก.6 บก.ทล."},
        {"id": "61", "sheet": "ส.ทล.1 (โคราช)", "fullName": "ส.ทล.1 กก.6 บก.ทล. (นครราชสีมา)"},
        {"id": "62", "sheet": "ส.ทล.2 (บุรีรัมย์)", "fullName": "ส.ทล.2 กก.6 บก.ทล. (บุรีรัมย์)"},
        {"id": "63", "sheet": "ส.ทล.3 (สุรินทร์)", "fullName": "ส.ทล.3 กก.6 บก.ทล. (สุรินทร์)"},
        {"id": "64", "sheet": "ส.ทล.4 (อุบลฯ)", "fullName": "ส.ทล.4 กก.6 บก.ทล. (อุบลราชธานี)"},
        {"id": "65", "sheet": "ส.ทล.5 (อำนาจเจริญ)", "fullName": "ส.ทล.5 กก.6 บก.ทล. (อำนาจเจริญ)"},
        {"id": "66", "sheet": "ส.ทล.6 (ชัยภูมิ)", "fullName": "ส.ทล.6 กก.6 บก.ทล. (ชัยภูมิ)"}
    ],
    7: [
        {"id": "70", "sheet": "ฝอ.กก.7", "fullName": "ฝอ.กก.7 บก.ทล."},
        {"id": "71", "sheet": "ส.ทล.1 (พังงา)", "fullName": "ส.ทล.1 กก.7 บก.ทล. (พังงา)"},
        {"id": "72", "sheet": "ส.ทล.2 (ตรัง)", "fullName": "ส.ทล.2 กก.7 บก.ทล. (ตรัง)"},
        {"id": "73", "sheet": "ส.ทล.3 (สงขลา)", "fullName": "ส.ทล.3 กก.7 บก.ทล. (สงขลา)"},
        {"id": "74", "sheet": "ส.ทล.4 (นครศรีฯ)", "fullName": "ส.ทล.4 กก.7 บก.ทล. (นครศรีธรรมราช)"},
        {"id": "75", "sheet": "ส.ทล.5 (ปัตตานี)", "fullName": "ส.ทล.5 กก.7 บก.ทล. (ปัตตานี)"}
    ],
    8: [
        {"id": "80", "sheet": "ฝอ.กก.8", "fullName": "ฝอ.กก.8 บก.ทล."},
        {"id": "81", "sheet": "ส.ทล.1 (มอเตอร์เวย์)", "fullName": "ส.ทล.1 กก.8 บก.ทล. (ชลบุรี มอเตอร์เวย์)"},
        {"id": "82", "sheet": "ส.ทล.2 (วงแหวน)", "fullName": "ส.ทล.2 กก.8 บก.ทล. (ปทุมธานี วงแหวน)"},
        {"id": "83", "sheet": "ส.ทล.3 (M6)", "fullName": "ส.ทล.3 กก.8 บก.ทล. (นครราชสีมา M6)"},
        {"id": "84", "sheet": "ส.ทล.4 (M81)", "fullName": "ส.ทล.4 กก.8 บก.ทล. (กาญจนบุรี M81)"}
    ]
}


def build_header_row_1_template(div_num: int, stations: list, output_dir: str):
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    # Styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    sample_font = Font(name="Tahoma", size=10, italic=True, color="555555")
    body_font = Font(name="Tahoma", size=10)
    title_font = Font(name="Tahoma", size=14, bold=True, color="1E3A8A")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # EXACT 11 COLUMNS STARTING AT ROW 1:
    headers = [
        "ลำดับ",
        "ยศ - ชื่อ - นามสกุล",
        "เบอร์โทรศัพท์",
        "กองกำกับการ (กก.)",
        "สถานีตำรวจทางหลวง (ส.ทล.)",
        "ตำแหน่ง / สิทธิ์ในระบบ",
        "สถานะการปฏิบัติงาน",
        "วันที่เริ่มช่วยราชการ",
        "วันที่สิ้นสุดช่วยราชการ",
        "ต้นทางที่ส่งมาช่วยราชการ",
        "ปลายทางที่ไปช่วยราชการ"
    ]

    for st in stations:
        ws = wb.create_sheet(title=st["sheet"])
        ws.views.sheetView[0].showGridLines = True

        # Check if this is the ฝอ. sheet
        is_hq_sheet = st["id"].endswith("0")
        station_col_val = f"กก.{div_num} บก.ทล." if is_hq_sheet else st["fullName"]

        # Row 1 IS THE HEADER ROW!
        ws.row_dimensions[1].height = 28
        for col_num, h_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = h_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 50 data rows per station sheet (Rows 2 to 51)
        for seq in range(1, 51):
            r = seq + 1
            ws.row_dimensions[r].height = 22

            if is_hq_sheet:
                role_val = "ฝอ.กก." if seq == 1 else ("ผกก." if seq == 2 else "ผู้ปฏิบัติประจำหน่วย")
            else:
                role_val = "สิบเวร / Admin สถานี" if seq == 1 else ("หัวหน้าหน่วยบริการ" if seq == 2 else "ผู้ปฏิบัติประจำหน่วย")

            row_values = [
                seq,
                "", # FullName
                "", # Phone
                f"กก.{div_num} บก.ทล.",
                station_col_val,
                role_val,
                "ปฏิบัติงานปกติ",
                "-",
                "-",
                "", # ต้นทางที่ส่งมาช่วยราชการ
                ""  # ปลายทางที่ไปช่วยราชการ
            ]

            for c_num, val in enumerate(row_values, 1):
                cell = ws.cell(row=r, column=c_num)
                cell.value = val
                cell.font = body_font if seq > 1 else sample_font
                cell.border = thin_border

                if c_num in [1, 3, 4, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                if c_num == 3:
                    cell.number_format = "@"

        # Role Data Validation
        dv_role = DataValidation(type="list", formula1='"ผู้ปฏิบัติประจำหน่วย,สิบเวร / Admin สถานี,หัวหน้าหน่วยบริการ,ฝอ.กก.,ผกก."', allow_blank=True)
        ws.add_data_validation(dv_role)
        dv_role.add("F2:F51")

        # Duty Status Data Validation
        dv_status = DataValidation(type="list", formula1='"ปฏิบัติงานปกติ,ไปช่วยราชการต่างหน่วย,มาช่วยราชการที่หน่วย,ลา / อบรม"', allow_blank=True)
        ws.add_data_validation(dv_status)
        dv_status.add("G2:G51")

        # Column Widths
        col_widths = {
            "A": 8,
            "B": 28,
            "C": 18,
            "D": 16,
            "E": 34,
            "F": 24,
            "G": 22,
            "H": 20,
            "I": 20,
            "J": 25,
            "K": 25
        }
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w

    # Add Guide Sheet at end
    ws_guide = wb.create_sheet(title="คำแนะนำการกรอกข้อมูล")
    ws_guide.views.sheetView[0].showGridLines = True
    ws_guide["A1"] = f"คำแนะนำการกรอกแบบฟอร์มข้อมูลเจ้าหน้าที่ - กก.{div_num} บก.ทล."
    ws_guide["A1"].font = title_font

    guides = [
        ("ชื่อคอลัมน์", "ความสำคัญ", "รายละเอียดและข้อแนะนำ"),
        ("ลำดับ", "อัตโนมัติ", "ลำดับที่ 1 - 50 ประจำสถานี"),
        ("ยศ - ชื่อ - นามสกุล", "บังคับ", "ระบุยศ ชื่อ และนามสกุลให้ครบถ้วน เช่น ด.ต. สมชาย สายตรวจ"),
        ("เบอร์โทรศัพท์", "บังคับ", "ระบุเบอร์โทรศัพท์มือถือ 10 หลัก (เช่น 0812345678) โดยไม่ต้องใส่ขีด"),
        ("กองกำกับการ (กก.)", "อัตโนมัติ", f"ระบบเติมค่า 'กก.{div_num} บก.ทล.' ให้อัตโนมัติ"),
        ("สถานีตำรวจทางหลวง (ส.ทล.)", "อัตโนมัติ", f"สำหรับแท็บ ฝอ. จะเติมค่า 'กก.{div_num} บก.ทล.' ให้อัตโนมัติ ส่วนแท็บสถานีจะเติมชื่อสถานีตามแท็บให้อัตโนมัติ"),
        ("ตำแหน่ง / สิทธิ์ในระบบ", "บังคับ", "กดเลือกสิทธิ์ในระบบ HWPD Next Gen:\n - ผู้ปฏิบัติประจำหน่วย: สายตรวจ/ผู้บันทึกรายงาน\n - สิบเวร / Admin สถานี: อนุมัติรายงานประจำสถานี\n - หัวหน้าหน่วยบริการ: หัวหน้าตู้/หน่วยบริการ\n - ฝอ.กก.: ฝ่ายอำนวยการระดับ กก.\n - ผกก.: ผู้กำกับการ"),
        ("สถานะการปฏิบัติงาน", "บังคับ", "กดเลือกสถานะปัจจุบัน:\n - ปฏิบัติงานปกติ\n - ไปช่วยราชการต่างหน่วย\n - มาช่วยราชการที่หน่วย\n - ลา / อบรม"),
        ("วันที่เริ่ม / สิ้นสุดช่วยราชการ", "กรณีช่วยราชการ", "ระบุวันที่ในรูปแบบ YYYY-MM-DD (เช่น 2026-08-01) หากไม่ได้ช่วยราชการให้ใส่ -"),
        ("ต้นทางที่ส่งมาช่วยราชการ", "กรณีมาช่วยราชการ", "ระบุชื่อหน่วยงานเดิมต้นทาง กรณีข้าราชการตำรวจมาช่วยราชการที่หน่วยนี้ (เช่น ส.ทล.2 กก.5)"),
        ("ปลายทางที่ไปช่วยราชการ", "กรณีไปช่วยราชการ", "ระบุชื่อหน่วยงานปลายทาง กรณีข้าราชการตำรวจของหน่วยนี้ไปช่วยราชการที่อื่น (เช่น บก.ปปป.)")
    ]

    ws_guide.row_dimensions[3].height = 26
    for c, h in enumerate(guides[0], 1):
        cell = ws_guide.cell(row=3, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row_data in enumerate(guides[1:], 4):
        ws_guide.row_dimensions[r_idx].height = 35
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_guide.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if c_idx == 1:
                cell.font = Font(name="Tahoma", size=10, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_guide.column_dimensions["A"].width = 30
    ws_guide.column_dimensions["B"].width = 18
    ws_guide.column_dimensions["C"].width = 75

    wb.remove(default_sheet)

    file_name = f"แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่_กก{div_num}_บก.ทล.xlsx"
    file_path = os.path.join(output_dir, file_name)
    wb.save(file_path)
    print(f"Generated Row 1 Header Workbook: {file_name}")


if __name__ == "__main__":
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    
    out_dir = os.path.join(base_dir, "แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่_แยกตาม_กก")
    os.makedirs(out_dir, exist_ok=True)

    for div_num, stations in STATIONS_DATA.items():
        build_header_row_1_template(div_num, stations, out_dir)

    print(f"\nSUCCESSFULLY UPDATED ALL 8 DIVISIONS EXCEL FILES (HEADER AT ROW 1) IN: {out_dir}")
