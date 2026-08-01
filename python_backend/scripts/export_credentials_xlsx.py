"""
รวมไฟล์รหัสผ่านจากสคริปต์สร้างบัญชี ออกเป็น Excel ไฟล์เดียวสำหรับแจกหน่วย

    python scripts/export_credentials_xlsx.py            # ใช้ไฟล์ CSV ล่าสุดอัตโนมัติ
    python scripts/export_credentials_xlsx.py --out ก.xlsx

ได้ชีต "รายชื่อบัญชีทั้งหมด" หนึ่งชีต แล้วแยกชีตราย กก. อีกชุด เพื่อส่งให้แต่ละกอง
โดยไม่ต้องกรองเอง และไม่ต้องเห็นรหัสของกองอื่น

จังหวัดกับชื่อสถานีอ่านจาก STATION_CONFIG ปัจจุบัน ไม่ได้ลอกจาก CSV เพราะไฟล์ CSV
ที่สร้างไว้ก่อนข้อมูลสถานีถูกพอร์ตเข้ามา ยังมีจังหวัดเป็นค่า placeholder ค้างอยู่

ไฟล์ผลลัพธ์มีรหัสผ่านแบบไม่เข้ารหัส ตั้งชื่อขึ้นต้น credentials_ ให้ตรงกับ .gitignore
แจกครบแล้วลบทิ้ง และควรให้เจ้าหน้าที่เปลี่ยนรหัสหลังล็อกอินครั้งแรก
"""

import argparse
import csv
import glob
import os
import sys
from collections import Counter, OrderedDict
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from app.core.config import get_station_data  # noqa: E402

ROLE_TH = {
    "Super_Commander": "ผู้บังคับการ",
    "HQ_Admin": "ฝ่ายอำนวยการ บก.",
    "Division_Commander": "ผู้กำกับการ",
    "Division_Admin": "ฝ่ายอำนวยการ กก.",
    "Station_Admin": "หัวหน้าสถานี",
    "สิบเวร": "สิบเวร",
    "Unit_Staff": "เจ้าหน้าที่ผู้ปฏิบัติ",
}

COLUMNS = ["Username", "Password", "ประเภทบัญชี", "สิทธิ์", "Role", "กก.", "Station_ID", "หน่วย/สถานี", "จังหวัด"]
WIDTHS = [14, 14, 18, 22, 20, 11, 12, 30, 24]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
# รหัสพวกนี้ต้องอ่านจากกระดาษแล้วพิมพ์ตาม ใช้ฟอนต์ที่แยก 0/O และ 1/l ออกจากกัน
MONO = Font(name="Consolas", size=11)
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def read_csv(path: str):
    with open(path, encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def merge_all(pattern: str):
    """
    รวมทุกไฟล์ที่เข้าคู่กับ pattern ไฟล์ใหม่ทับไฟล์เก่าเมื่อ Username ซ้ำ

    ห้ามหยิบแค่ไฟล์ล่าสุด เพราะ create_station_users.py --apply เขียนเฉพาะบัญชีที่
    "เพิ่งสร้าง" ลงไฟล์ใหม่ ไม่ใช่ทั้งชุด ตอนเพิ่ม fo0 ทีหลังจึงได้ไฟล์ที่มีบรรทัดเดียว
    ถ้าอ่านไฟล์นั้นไฟล์เดียวก็จะเหลือบัญชีเดียวจาก 53
    """
    merged = OrderedDict()
    files = sorted(glob.glob(pattern))
    for path in files:
        for row in read_csv(path):
            username = str(row.get("Username", "")).strip()
            if username:
                merged[username] = row
    return list(merged.values()), files


def build_records(unit_rows, operator_rows):
    records = []

    def add(row, account_type):
        station = str(row.get("Station_ID", "")).strip()
        data = get_station_data(station) or {}
        records.append(
            {
                "Username": row.get("Username", ""),
                "Password": row.get("Password", ""),
                "ประเภทบัญชี": account_type,
                "สิทธิ์": ROLE_TH.get(row.get("Role", ""), row.get("Role", "")),
                "Role": row.get("Role", ""),
                "กก.": "ส่วนกลาง" if station == "00" else f"กก.{station[:1]}",
                "Station_ID": station,
                "หน่วย/สถานี": data.get("fullName") or row.get("หน่วย") or row.get("สถานี", ""),
                "จังหวัด": data.get("province", ""),
            }
        )

    for row in unit_rows:
        add(row, "บัญชีประจำหน่วย")
    for row in operator_rows:
        add(row, "บัญชีผู้ปฏิบัติ")

    records.sort(key=lambda r: (0 if r["Station_ID"] == "00" else int(r["Station_ID"][:1] or 0),
                                r["Station_ID"], r["Username"]))
    return records


def write_sheet(worksheet, records) -> None:
    worksheet.append(COLUMNS)
    for index, _ in enumerate(COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=index)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[get_column_letter(index)].width = WIDTHS[index - 1]
    worksheet.row_dimensions[1].height = 24

    for line, record in enumerate(records, start=2):
        worksheet.append([record[column] for column in COLUMNS])
        for index in range(1, len(COLUMNS) + 1):
            cell = worksheet.cell(row=line, column=index)
            cell.border = BORDER
            if line % 2 == 0:
                cell.fill = BAND_FILL
        worksheet.cell(row=line, column=1).font = MONO
        worksheet.cell(row=line, column=2).font = MONO

    worksheet.freeze_panes = "A2"
    if records:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(records) + 1}"


def main() -> int:
    parser = argparse.ArgumentParser(description="รวมไฟล์รหัสผ่านออกเป็น Excel สำหรับแจกหน่วย")
    parser.add_argument("--unit", default="", help="CSV บัญชีประจำหน่วย (ค่าเริ่มต้น: ไฟล์ล่าสุด)")
    parser.add_argument("--operators", default="", help="CSV บัญชีผู้ปฏิบัติ (ค่าเริ่มต้น: ไฟล์ล่าสุด)")
    parser.add_argument("--out", default="", help="ชื่อไฟล์ผลลัพธ์")
    args = parser.parse_args()

    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if args.unit:
        unit_rows, unit_files = read_csv(args.unit), [args.unit]
    else:
        unit_rows, unit_files = merge_all(os.path.join(root, "credentials_2*.csv"))
    if args.operators:
        op_rows, op_files = read_csv(args.operators), [args.operators]
    else:
        op_rows, op_files = merge_all(os.path.join(root, "credentials_operators_*.csv"))

    if not unit_rows and not op_rows:
        print("ไม่พบไฟล์ credentials_*.csv — รัน create_station_users.py / create_operator_users.py ก่อน")
        return 1

    print(f"บัญชีประจำหน่วย : {len(unit_rows)} บัญชี จาก {len(unit_files)} ไฟล์")
    for path in unit_files:
        print(f"    {os.path.basename(path)}")
    print(f"บัญชีผู้ปฏิบัติ  : {len(op_rows)} บัญชี จาก {len(op_files)} ไฟล์")
    for path in op_files:
        print(f"    {os.path.basename(path)}")
    print()

    records = build_records(unit_rows, op_rows)
    if not records:
        print("ไฟล์ CSV ว่างเปล่า")
        return 1

    workbook = Workbook()
    write_sheet(workbook.active, records)
    workbook.active.title = "รายชื่อบัญชีทั้งหมด"

    # แยกชีตราย กก. เรียงตามลำดับที่เจอในข้อมูล ซึ่งเรียงมาแล้วตั้งแต่ build_records
    groups = OrderedDict()
    for record in records:
        groups.setdefault(record["กก."], []).append(record)
    for name, rows in groups.items():
        write_sheet(workbook.create_sheet(title=name[:31]), rows)

    out = args.out or os.path.join(root, f"credentials_hwpd_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    workbook.save(out)

    print(f"สร้างไฟล์แล้ว: {out}")
    print(f"รวม {len(records)} บัญชี | {len(groups) + 1} ชีต")
    for kind, count in Counter(r["ประเภทบัญชี"] for r in records).items():
        print(f"  {kind}: {count}")
    print()
    for name, rows in groups.items():
        print(f"  ชีต {name:<10} {len(rows)} บัญชี")
    print("\nไฟล์นี้มีรหัสผ่านแบบไม่เข้ารหัส ถูก gitignore ไว้ แจกครบแล้วลบทิ้ง")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
