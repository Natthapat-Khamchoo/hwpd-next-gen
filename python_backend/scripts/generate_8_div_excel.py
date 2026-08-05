import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Config for 8 Divisions
DIVISIONS_CONFIG = {
    1: {
        "name": "กก.1 บก.ทล.",
        "stations": [
            "ฝอ.กก.1 บก.ทล.",
            "ส.ทล.1 กก.1 บก.ทล. (อยุธยา)",
            "ส.ทล.2 กก.1 บก.ทล. (สระบุรี)",
            "ส.ทล.3 กก.1 บก.ทล. (ลพบุรี)",
            "ส.ทล.4 กก.1 บก.ทล. (นครสวรรค์)",
            "ส.ทล.5 กก.1 บก.ทล. (เพชรบูรณ์)",
            "ส.ทล.6 กก.1 บก.ทล. (สิงห์บุรี)"
        ],
        "sample_station": "ส.ทล.1 กก.1 บก.ทล. (อยุธยา)",
        "sample_unit": "หน่วยบริการวังน้อย"
    },
    2: {
        "name": "กก.2 บก.ทล.",
        "stations": [
            "ฝอ.กก.2 บก.ทล.",
            "ส.ทล.1 กก.2 บก.ทล. (นครปฐม)",
            "ส.ทล.2 กก.2 บก.ทล. (เพชรบุรี)",
            "ส.ทล.3 กก.2 บก.ทล. (ประจวบคีรีขันธ์)",
            "ส.ทล.4 กก.2 บก.ทล. (ชุมพร)",
            "ส.ทล.5 กก.2 บก.ทล. (สุราษฎร์ธานี)",
            "ส.ทล.6 กก.2 บก.ทล. (สุพรรณบุรี)"
        ],
        "sample_station": "ส.ทล.1 กก.2 บก.ทล. (นครปฐม)",
        "sample_unit": "หน่วยบริการท่าตำหนัก"
    },
    3: {
        "name": "กก.3 บก.ทล.",
        "stations": [
            "ฝอ.กก.3 บก.ทล.",
            "ส.ทล.1 กก.3 บก.ทล. (ฉะเชิงเทรา)",
            "ส.ทล.2 กก.3 บก.ทล. (ชลบุรี)",
            "ส.ทล.3 กก.3 บก.ทล. (ระยอง)",
            "ส.ทล.4 กก.3 บก.ทล. (จันทบุรี)",
            "ส.ทล.5 กก.3 บก.ทล. (ปราจีนบุรี)"
        ],
        "sample_station": "ส.ทล.1 กก.3 บก.ทล. (ฉะเชิงเทรา)",
        "sample_unit": "หน่วยบริการบางปะกง"
    },
    4: {
        "name": "กก.4 บก.ทล.",
        "stations": [
            "ฝอ.กก.4 บก.ทล.",
            "ส.ทล.1 กก.4 บก.ทล. (ร้อยเอ็ด)",
            "ส.ทล.2 กก.4 บก.ทล. (ขอนแก่น)",
            "ส.ทล.3 กก.4 บก.ทล. (อุดรธานี)",
            "ส.ทล.4 กก.4 บก.ทล. (เลย)",
            "ส.ทล.5 กก.4 บก.ทล. (สกลนคร)"
        ],
        "sample_station": "ส.ทล.2 กก.4 บก.ทล. (ขอนแก่น)",
        "sample_unit": "หน่วยบริการน้ำพอง"
    },
    5: {
        "name": "กก.5 บก.ทล.",
        "stations": [
            "ฝอ.กก.5 บก.ทล.",
            "ส.ทล.1 กก.5 บก.ทล. (ตาก)",
            "ส.ทล.2 กก.5 บก.ทล. (ลำปาง)",
            "ส.ทล.3 กก.5 บก.ทล. (พิษณุโลก)",
            "ส.ทล.4 กก.5 บก.ทล. (เชียงใหม่)",
            "ส.ทล.5 กก.5 บก.ทล. (พะเยา)",
            "ส.ทล.6 กก.5 บก.ทล. (แพร่)"
        ],
        "sample_station": "ส.ทล.4 กก.5 บก.ทล. (เชียงใหม่)",
        "sample_unit": "หน่วยบริการดอนแก้ว"
    },
    6: {
        "name": "กก.6 บก.ทล.",
        "stations": [
            "ฝอ.กก.6 บก.ทล.",
            "ส.ทล.1 กก.6 บก.ทล. (นครราชสีมา)",
            "ส.ทล.2 กก.6 บก.ทล. (บุรีรัมย์)",
            "ส.ทล.3 กก.6 บก.ทล. (สุรินทร์)",
            "ส.ทล.4 กก.6 บก.ทล. (อุบลราชธานี)",
            "ส.ทล.5 กก.6 บก.ทล. (อำนาจเจริญ)",
            "ส.ทล.6 กก.6 บก.ทล. (ชัยภูมิ)"
        ],
        "sample_station": "ส.ทล.1 กก.6 บก.ทล. (นครราชสีมา)",
        "sample_unit": "หน่วยบริการกลางดง"
    },
    7: {
        "name": "กก.7 บก.ทล.",
        "stations": [
            "ฝอ.กก.7 บก.ทล.",
            "ส.ทล.1 กก.7 บก.ทล. (พังงา)",
            "ส.ทล.2 กก.7 บก.ทล. (ตรัง)",
            "ส.ทล.3 กก.7 บก.ทล. (สงขลา)",
            "ส.ทล.4 กก.7 บก.ทล. (นครศรีธรรมราช)",
            "ส.ทล.5 กก.7 บก.ทล. (ปัตตานี)"
        ],
        "sample_station": "ส.ทล.3 กก.7 บก.ทล. (สงขลา)",
        "sample_unit": "หน่วยบริการทุ่งลุง"
    },
    8: {
        "name": "กก.8 บก.ทล.",
        "stations": [
            "ฝอ.กก.8 บก.ทล.",
            "ส.ทล.1 กก.8 บก.ทล. (ชลบุรี มอเตอร์เวย์)",
            "ส.ทล.2 กก.8 บก.ทล. (ปทุมธานี วงแหวน)",
            "ส.ทล.3 กก.8 บก.ทล. (นครราชสีมา M6)",
            "ส.ทล.4 กก.8 บก.ทล. (กาญจนบุรี M81)"
        ],
        "sample_station": "ส.ทล.1 กก.8 บก.ทล. (ชลบุรี มอเตอร์เวย์)",
        "sample_unit": "หน่วยบริการพานทอง"
    }
}


def create_excel_for_division(div_num: int, info: dict, output_dir: str):
    wb = openpyxl.Workbook()

    # Sheet 1: Form
    ws = wb.active
    ws.title = f"แบบฟอร์มเจ้าหน้าที่ กก.{div_num}"
    ws.views.sheetView[0].showGridLines = True

    # Styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    sample_font = Font(name="Tahoma", size=10, italic=True, color="555555")
    body_font = Font(name="Tahoma", size=10)
    title_font = Font(name="Tahoma", size=14, bold=True, color="1E3A8A")
    note_font = Font(name="Tahoma", size=10, italic=True, color="DC2626")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # Title block
    ws.merge_cells("A1:L1")
    ws["A1"] = f"แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่ตำรวจทางหลวง - {info['name']}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = f"* สำหรับเจ้าหน้าที่ในสังกัด {info['name']} | กรุณาอย่าลบหรือแก้ไขชื่อหัวตาราง และใช้เมนู Dropdown ในช่องที่มีตัวเลือก"
    ws["A2"].font = note_font
    ws["A2"].alignment = Alignment(vertical="center")

    # Headers
    headers = [
        "ลำดับ",
        "ยศ - ชื่อ - นามสกุล",
        "เลขประจำตัวตำรวจ / รหัสกำลังพล",
        "เบอร์โทรศัพท์",
        "กองกำกับการ (กก.)",
        "สถานีตำรวจทางหลวง (ส.ทล.)",
        "หน่วยบริการ / ตู้สายตรวจ",
        "ตำแหน่ง / สิทธิ์ในระบบ",
        "สถานะการปฏิบัติงาน",
        "วันที่เริ่มช่วยราชการ",
        "วันที่สิ้นสุดช่วยราชการ",
        "หมายเหตุ"
    ]

    ws.row_dimensions[4].height = 28

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Sample Row (Row 5)
    sample_row = [
        1,
        "ด.ต. สมชาย สายตรวจ (ตัวอย่าง)",
        "65012345",
        "0812345678",
        info["name"],
        info["sample_station"],
        info["sample_unit"],
        "ผู้ปฏิบัติประจำหน่วย",
        "ปฏิบัติงานปกติ",
        "-",
        "-",
        "ตัวอย่างการกรอกข้อมูล"
    ]

    ws.row_dimensions[5].height = 24
    for col_num, val in enumerate(sample_row, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = val
        cell.font = sample_font
        cell.border = thin_border
        if col_num in [1, 3, 4, 9, 10, 11]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_num == 4:
            cell.number_format = "@"

    # Blank rows for user filling (Row 6 to 150)
    for r in range(6, 151):
        ws.row_dimensions[r].height = 22
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = thin_border
            if c == 1:
                cell.value = r - 5
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 5:
                cell.value = info["name"]  # Pre-fill division name
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [3, 4, 9, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if c == 4:
                cell.number_format = "@"

    # Data Validation for Stations
    stations_str = ",".join(info["stations"])
    dv_station = DataValidation(
        type="list",
        formula1=f'"{stations_str}"',
        allow_blank=True
    )
    dv_station.error = "กรุณาเลือก สถานีตำรวจทางหลวง จากรายการที่มีให้"
    dv_station.errorTitle = "ข้อมูลไม่ถูกต้อง"
    ws.add_data_validation(dv_station)
    dv_station.add("F5:F150")

    # Data Validation for Roles
    dv_role = DataValidation(
        type="list",
        formula1='"ผู้ปฏิบัติประจำหน่วย,สิบเวร / Admin สถานี,หัวหน้าหน่วยบริการ,ฝอ.กก.,ผกก."',
        allow_blank=True
    )
    dv_role.error = "กรุณาเลือก ตำแหน่ง/สิทธิ์ จากรายการที่มีให้"
    dv_role.errorTitle = "ข้อมูลไม่ถูกต้อง"
    ws.add_data_validation(dv_role)
    dv_role.add("H5:H150")

    # Data Validation for Duty Status
    dv_status = DataValidation(
        type="list",
        formula1='"ปฏิบัติงานปกติ,ไปช่วยราชการต่างหน่วย,มาช่วยราชการที่หน่วย,ลา / อบรม"',
        allow_blank=True
    )
    dv_status.error = "กรุณาเลือก สถานะการปฏิบัติงาน จากรายการที่มีให้"
    dv_status.errorTitle = "ข้อมูลไม่ถูกต้อง"
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I150")

    # Adjust Column Widths
    col_widths = {
        "A": 8,
        "B": 28,
        "C": 22,
        "D": 18,
        "E": 16,
        "F": 34,
        "G": 24,
        "H": 24,
        "I": 22,
        "J": 20,
        "K": 20,
        "L": 25
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Sheet 2: Guide
    ws_guide = wb.create_sheet(title="คำแนะนำการกรอกข้อมูล")
    ws_guide.views.sheetView[0].showGridLines = True

    ws_guide["A1"] = f"คำแนะนำการกรอกแบบฟอร์มข้อมูลเจ้าหน้าที่ - {info['name']}"
    ws_guide["A1"].font = title_font

    guides = [
        ("ชื่อคอลัมน์", "ความสำคัญ", "รายละเอียดและข้อแนะนำ"),
        ("ยศ - ชื่อ - นามสกุล", "บังคับ", "ระบุยศ ชื่อ และนามสกุลให้ครบถ้วน เช่น ด.ต. สมชาย สายตรวจ"),
        ("เลขประจำตัวตำรวจ / รหัสกำลังพล", "บังคับ", "ระบุรหัสกำลังพล หรือเลขประจำตัวตำรวจ (ตัวเลขเท่านั้น)"),
        ("เบอร์โทรศัพท์", "บังคับ", "ระบุเบอร์โทรศัพท์มือถือ 10 หลัก (เช่น 0812345678) โดยไม่ต้องใส่ขีด"),
        ("กองกำกับการ (กก.)", "บังคับ", f"ระบบเติมค่า '{info['name']}' ให้อัตโนมัติ"),
        ("สถานีตำรวจทางหลวง (ส.ทล.)", "บังคับ", f"กดเลือกสถานีสังกัดใน {info['name']} จากเมนู Dropdown"),
        ("หน่วยบริการ / ตู้สายตรวจ", "บังคับ", "ระบุหน่วยบริการประจำ เช่น หน่วยบริการดอนแก้ว (ถ้าประจำสถานีให้ระบุชื่อสถานี)"),
        ("ตำแหน่ง / สิทธิ์ในระบบ", "บังคับ", "กดเลือกสิทธิ์ในระบบ HWPD Next Gen:\n - ผู้ปฏิบัติประจำหน่วย: สายตรวจ/ผู้บันทึก\n - สิบเวร / Admin สถานี: อนุมัติรายงานประจำสถานี\n - หัวหน้าหน่วยบริการ: หัวหน้าตู้/หน่วยบริการ\n - ฝอ.กก.: ฝ่ายอำนวยการระดับ กก.\n - ผกก.: ผู้กำกับการ"),
        ("สถานะการปฏิบัติงาน", "บังคับ", "กดเลือกสถานะปัจจุบัน:\n - ปฏิบัติงานปกติ\n - ไปช่วยราชการต่างหน่วย\n - มาช่วยราชการที่หน่วย\n - ลา / อบรม"),
        ("วันที่เริ่ม / สิ้นสุดช่วยราชการ", "กรณีช่วยราชการ", "ระบุวันที่ในรูปแบบ YYYY-MM-DD (เช่น 2026-08-01) หากไม่ได้ช่วยราชการให้ละเว้นไว้"),
        ("หมายเหตุ", "ตัวเลือก", "ระบุข้อมูลเพิ่มเติม เช่น ช่วยราชการมาจาก ส.ทล.2 เป็นต้น")
    ]

    ws_guide.row_dimensions[3].height = 26
    for c, h in enumerate(guides[0], 1):
        cell = ws_guide.cell(row=3, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row_data in enumerate(guides[1:], 4):
        ws_guide.row_dimensions[r_idx].height = 35
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_guide.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if c_idx == 1:
                cell.font = Font(name="Tahoma", size=10, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_guide.column_dimensions["A"].width = 30
    ws_guide.column_dimensions["B"].width = 18
    ws_guide.column_dimensions["C"].width = 75

    file_name = f"แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่_กก{div_num}_บก.ทล.xlsx"
    file_path = os.path.join(output_dir, file_name)
    wb.save(file_path)
    print(f"Generated: {file_name}")


if __name__ == "__main__":
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    out_dir = os.path.join(base_dir, "แบบฟอร์มแยกตาม_กก")
    os.makedirs(out_dir, exist_ok=True)

    for div_num, info in DIVISIONS_CONFIG.items():
        create_excel_for_division(div_num, info, out_dir)

    print(f"\nALL 8 DIVISIONS EXCEL FILES CREATED IN: {out_dir}")
