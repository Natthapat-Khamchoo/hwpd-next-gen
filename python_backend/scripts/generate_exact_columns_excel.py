# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

STATIONS_DATA = {
    1: [
        {"id": "10", "sheet": "ฝอ.กก.1", "fullName": "ฝอ.กก.1 บก.ทล.", "units": ["ฝอ.กก.1"]},
        {"id": "11", "sheet": "ส.ทล.1 (อยุธยา)", "fullName": "ส.ทล.1 กก.1 บก.ทล. (อยุธยา)", "units": ["วังน้อย", "รังสิต", "เอเซีย", "บางปะอิน", "นพวงศ์", "ราชพฤกษ์"]},
        {"id": "12", "sheet": "ส.ทล.2 (สระบุรี)", "fullName": "ส.ทล.2 กก.1 บก.ทล. (สระบุรี)", "units": ["พหลโยธิน เขต 1", "พหลโยธิน เขต 2", "เฉลิมพระเกียรติ", "ทับกวาง", "มิตรภาพ", "บ้านนา", "บางอ้อ"]},
        {"id": "13", "sheet": "ส.ทล.3 (ลพบุรี)", "fullName": "ส.ทล.3 กก.1 บก.ทล. (ลพบุรี)", "units": ["เมืองลพบุรี", "โคกสำโรง", "ชัยบาดาล", "พัฒนานิคม"]},
        {"id": "14", "sheet": "ส.ทล.4 (นครสวรรค์)", "fullName": "ส.ทล.4 กก.1 บก.ทล. (นครสวรรค์)", "units": ["เขาหน่อ", "ราชรถ (เก้าเลี้ยว)", "ท่าน้ำอ้อย", "ท่าตะโก", "หนองบัว", "อุทัยธานี"]},
        {"id": "15", "sheet": "ส.ทล.5 (เพชรบูรณ์)", "fullName": "ส.ทล.5 กก.1 บก.ทล. (เพชรบูรณ์)", "units": ["ศรีเทพ", "ราหุล", "หล่มสัก", "เขาทราย", "วชิรบารมี"]},
        {"id": "16", "sheet": "ส.ทล.6 (สิงห์บุรี)", "fullName": "ส.ทล.6 กก.1 บก.ทล. (สิงห์บุรี)", "units": ["อ่างทอง", "อินทร์บุรี", "สรรคบุรี", "ชัยนาท"]}
    ],
    2: [
        {"id": "20", "sheet": "ฝอ.กก.2", "fullName": "ฝอ.กก.2 บก.ทล.", "units": ["ฝอ.กก.2"]},
        {"id": "21", "sheet": "ส.ทล.1 (นครปฐม)", "fullName": "ส.ทล.1 กก.2 บก.ทล. (นครปฐม)", "units": ["สมุทรสงคราม", "มหาชัย", "ศรีสำราญ", "สาย 4", "ท่าตำหนัก", "หนองดินแดง", "ทัพหลวง", "กำแพงแสน", "บางเลน", "ดอนยายหอม", "บางโทรัด"]},
        {"id": "22", "sheet": "ส.ทล.2 (เพชรบุรี)", "fullName": "ส.ทล.2 กก.2 บก.ทล. (เพชรบุรี)", "units": ["บ้านโป่ง", "โพธาราม", "วังมะนาว", "เขตตรวจที่ 4", "ชะอำ"]},
        {"id": "23", "sheet": "ส.ทล.3 (ประจวบฯ)", "fullName": "ส.ทล.3 กก.2 บก.ทล. (ประจวบคีรีขันธ์)", "units": ["หัวหิน", "ห้วยมงคล", "กุยบุรี", "เกาะหลัก", "ทับสะเเก", "บางสะพาน", "ไชยราช"]},
        {"id": "24", "sheet": "ส.ทล.4 (ชุมพร)", "fullName": "ส.ทล.4 กก.2 บก.ทล. (ชุมพร)", "units": ["ท่าแซะ", "หลังสวน", "ทับหลี", "ระนอง"]},
        {"id": "25", "sheet": "ส.ทล.5 (สุราษฎร์ฯ)", "fullName": "ส.ทล.5 กก.2 บก.ทล. (สุราษฎร์ธานี)", "units": ["ไชยา", "เวียงสระ", "กาญจนดิษฐ์", "ท่าโรงช้าง"]},
        {"id": "26", "sheet": "ส.ทล.6 (สุพรรณบุรี)", "fullName": "ส.ทล.6 กก.2 บก.ทล. (สุพรรณบุรี)", "units": ["ท่ามะกา", "ไทรโยค", "พระแท่น", "อู่ทอง", "สาลี", "ศรีประจันต์", "ด่านช้าง"]}
    ],
    3: [
        {"id": "30", "sheet": "ฝอ.กก.3", "fullName": "ฝอ.กก.3 บก.ทล.", "units": ["ฝอ.กก.3"]},
        {"id": "31", "sheet": "ส.ทล.1 (ฉะเชิงเทรา)", "fullName": "ส.ทล.1 กก.3 บก.ทล. (ฉะเชิงเทรา)", "units": ["บางพลี", "บางปู", "เขตตรวจที่ 3 (บางปะกง)", "แปลงยาว", "บางคล้า", "สุวินทวงศ์", "พระประแดง"]},
        {"id": "32", "sheet": "ส.ทล.2 (ชลบุรี)", "fullName": "ส.ทล.2 กก.3 บก.ทล. (ชลบุรี)", "units": ["สัตหีบ", "เขาคันทรง", "หนองใหญ่", "เขตตรวจที่ 6 (พนัสนิคม)"]},
        {"id": "33", "sheet": "ส.ทล.3 (ระยอง)", "fullName": "ส.ทล.3 กก.3 บก.ทล. (ระยอง)", "units": ["มาบข่า", "วังจันทร์"]},
        {"id": "34", "sheet": "ส.ทล.4 (จันทบุรี)", "fullName": "ส.ทล.4 กก.3 บก.ทล. (จันทบุรี)", "units": ["นายายอาม", "โป่งน้ำร้อน", "ขลุง"]},
        {"id": "35", "sheet": "ส.ทล.5 (ปราจีนบุรี)", "fullName": "ส.ทล.5 กก.3 บก.ทล. (ปราจีนบุรี)", "units": ["กบินบุรี", "สระแก้ว", "ศรีมหาโพธิ", "นาดี", "วังทอง"]}
    ],
    4: [
        {"id": "40", "sheet": "ฝอ.กก.4", "fullName": "ฝอ.กก.4 บก.ทล.", "units": ["ฝอ.กก.4"]},
        {"id": "41", "sheet": "ส.ทล.1 (ร้อยเอ็ด)", "fullName": "ส.ทล.1 กก.4 บก.ทล. (ร้อยเอ็ด)", "units": ["โพนทอง", "รอบเมือง", "สุวรรณภูมิ", "พยัคฆภูมิพิสัย", "บรบือ", "ยางตลาด", "สมเด็จ"]},
        {"id": "42", "sheet": "ส.ทล.2 (ขอนแก่น)", "fullName": "ส.ทล.2 กก.4 บก.ทล. (ขอนแก่น)", "units": ["พล", "บ้านไผ่", "น้ำพอง", "หนองเรือ", "ชุมแพ", "รอบเมือง"]},
        {"id": "43", "sheet": "ส.ทล.3 (อุดรธานี)", "fullName": "ส.ทล.3 กก.4 บก.ทล. (อุดรธานี)", "units": ["รอบเมือง", "โนนสะอาด", "หนองหาน", "หนองคาย", "บึงกาฬ"]},
        {"id": "44", "sheet": "ส.ทล.4 (เลย)", "fullName": "ส.ทล.4 กก.4 บก.ทล. (เลย)", "units": ["วังสะพุง", "หนองบัวลำภู", "ภูเรือ", "เชียงคาน"]},
        {"id": "45", "sheet": "ส.ทล.5 (สกลนคร)", "fullName": "ส.ทล.5 กก.4 บก.ทล. (สกลนคร)", "units": ["สร้างค้อ", "พังโคน", "นครพนม"]}
    ],
    5: [
        {"id": "50", "sheet": "ฝอ.กก.5", "fullName": "ฝอ.กก.5 บก.ทล.", "units": ["ฝอ.กก.5"]},
        {"id": "51", "sheet": "ส.ทล.1 (ตาก)", "fullName": "ส.ทล.1 กก.5 บก.ทล. (ตาก)", "units": ["คลองขลุง", "นครชุม", "สามเงา", "แม่สอด", "เขตตรวจ 3"]},
        {"id": "52", "sheet": "ส.ทล.2 (ลำปาง)", "fullName": "ส.ทล.2 กก.5 บก.ทล. (ลำปาง)", "units": ["เถิน", "เกาะคา", "ห้างฉัตร", "งาว", "ลำพูน", "ลี้"]},
        {"id": "53", "sheet": "ส.ทล.3 (พิษณุโลก)", "fullName": "ส.ทล.3 กก.5 บก.ทล. (พิษณุโลก)", "units": ["บ้านป่า", "สีหราชเดโชชัย", "ทรัพย์ไพรวัลย์", "สุโขทัย", "ศรีสัชนาลัย"]},
        {"id": "54", "sheet": "ส.ทล.4 (เชียงใหม่)", "fullName": "ส.ทล.4 กก.5 บก.ทล. (เชียงใหม่)", "units": ["สารภี", "ฮอด", "แม่สะเรียง", "ดอยสะเก็ด", "แม่แตง", "เขตครวจฝาง"]},
        {"id": "55", "sheet": "ส.ทล.5 (พะเยา)", "fullName": "ส.ทล.5 กก.5 บก.ทล. (พะเยา)", "units": ["แม่กา", "แม่อิง", "เชียงราย", "แม่จัน", "เชียงของ", "เขตตรวจ เวียงป่าเป้า"]},
        {"id": "56", "sheet": "ส.ทล.6 (แพร่)", "fullName": "ส.ทล.6 กก.5 บก.ทล. (แพร่)", "units": ["อุตรดิตถ์", "เด่นชัย", "แม่คำมี", "น่าน"]}
    ],
    6: [
        {"id": "60", "sheet": "ฝอ.กก.6", "fullName": "ฝอ.กก.6 บก.ทล.", "units": ["ฝอ.กก.6"]},
        {"id": "61", "sheet": "ส.ทล.1 (โคราช)", "fullName": "ส.ทล.1 กก.6 บก.ทล. (นครราชสีมา)", "units": ["กลางดง", "คลองไผ่", "สูงเนิน", "จอหอ", "บ้านส้ม", "สีดา", "ด่านขุนทด", "ปักธงชัย", "วังน้ำเขียว", "โชคชัย", "ห้วยแถลง", "ประทาย", "หนองบุญมาก"]},
        {"id": "62", "sheet": "ส.ทล.2 (บุรีรัมย์)", "fullName": "ส.ทล.2 กก.6 บก.ทล. (บุรีรัมย์)", "units": ["โนนดินแดง", "นางรอง", "ประโคนชัย", "พุทไธสง", "เขตตรวจด่านชั่ง (หนองกี่)", "เขตตรวจรอบเมือง"]},
        {"id": "63", "sheet": "ส.ทล.3 (สุรินทร์)", "fullName": "ส.ทล.3 กก.6 บก.ทล. (สุรินทร์)", "units": ["สังขะ", "ปราสาท", "ท่าตูม", "ศีขรภูมิ", "เขตตรวจรอบเมือง"]},
        {"id": "64", "sheet": "ส.ทล.4 (อุบลฯ)", "fullName": "ส.ทล.4 กก.6 บก.ทล. (อุบลราชธานี)", "units": ["หนองไผ่", "เดชอุดม", "พิบูลมังสาหาร", "เขมราฐ", "เขื่องใน", "กันทรลักษ์", "อุทุมพร"]},
        {"id": "65", "sheet": "ส.ทล.5 (อำนาจเจริญ)", "fullName": "ส.ทล.5 กก.6 บก.ทล. (อำนาจเจริญ)", "units": ["ยโสธร", "มุกดาหาร", "นิคมคำสร้อย", "เขตตรวจปทุมราชวงศา"]},
        {"id": "66", "sheet": "ส.ทล.6 (ชัยภูมิ)", "fullName": "ส.ทล.6 กก.6 บก.ทล. (ชัยภูมิ)", "units": ["คอนสาร", "ช่องสามหมอ", "ชัยภูมิ", "หนองบัวโคก", "เทพสถิต"]}
    ],
    7: [
        {"id": "70", "sheet": "ฝอ.กก.7", "fullName": "ฝอ.กก.7 บก.ทล.", "units": ["ฝอ.กก.7"]},
        {"id": "71", "sheet": "ส.ทล.1 (พังงา)", "fullName": "ส.ทล.1 กก.7 บก.ทล. (พังงา)", "units": ["กระบี่", "อ่าวลึก", "ตะกั่วป่า", "ภูเก็ต", "เขตตรวจที่ 1 (พังงา)"]},
        {"id": "72", "sheet": "ส.ทล.2 (ตรัง)", "fullName": "ส.ทล.2 กก.7 บก.ทล. (ตรัง)", "units": ["เขตตรวจที่ 1", "ปะเหลียน", "ห้วยยอด", "พัทลุง", "ป่าบอน"]},
        {"id": "73", "sheet": "ส.ทล.3 (สงขลา)", "fullName": "ส.ทล.3 กก.7 บก.ทล. (สงขลา)", "units": ["สิงหนคร", "จะนะ", "ทุ่งลุง", "รัตภูมิ", "สตูล"]},
        {"id": "74", "sheet": "ส.ทล.4 (นครศรีฯ)", "fullName": "ส.ทล.4 กก.7 บก.ทล. (นครศรีธรรมราช)", "units": ["หัวไทร", "สิชล", "ชะอวด", "ทุ่งสง"]},
        {"id": "75", "sheet": "ส.ทล.5 (ปัตตานี)", "fullName": "ส.ทล.5 กก.7 บก.ทล. (ปัตตานี)", "units": ["บ่อทอง", "กลาพอ", "หน่วยกู้ภัยฯ นราธิวาส", "เขตตรวจยะลา"]}
    ],
    8: [
        {"id": "80", "sheet": "ฝอ.กก.8", "fullName": "ฝอ.กก.8 บก.ทล.", "units": ["ฝอ.กก.8"]},
        {"id": "81", "sheet": "ส.ทล.1 (มอเตอร์เวย์)", "fullName": "ส.ทล.1 กก.8 บก.ทล. (ชลบุรี มอเตอร์เวย์)", "units": ["ลาดกระบัง", "พานทอง", "เขาเขียว", "บางละมุง", "มาบประชัน"]},
        {"id": "82", "sheet": "ส.ทล.2 (วงแหวน)", "fullName": "ส.ทล.2 กก.8 บก.ทล. (ปทุมธานี วงแหวน)", "units": ["รีพัฒน์", "ธัญบุรี", "ทับช้าง", "บางนา", "บางแค", "บางหญ่", "คูบางหลวง"]},
        {"id": "83", "sheet": "ส.ทล.3 (M6)", "fullName": "ส.ทล.3 กก.8 บก.ทล. (นครราชสีมา M6)", "units": ["มอเตอร์เวย์ M6(ลำตะคอง)"]},
        {"id": "84", "sheet": "ส.ทล.4 (M81)", "fullName": "ส.ทล.4 กก.8 บก.ทล. (กาญจนบุรี M81)", "units": ["บางแม่นาง", "ท่าม่วง"]}
    ]
}


def build_exact_division_workbook(div_num: int, stations: list, output_dir: str):
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    # Styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    sample_font = Font(name="Tahoma", size=10, italic=True, color="555555")
    body_font = Font(name="Tahoma", size=10)
    title_font = Font(name="Tahoma", size=14, bold=True, color="1E3A8A")
    note_font = Font(name="Tahoma", size=10, italic=True, color="DC2626")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # EXACT 9 COLUMNS requested by user:
    # 1. ลำดับ
    # 2. ยศ - ชื่อ - นามสกุล
    # 3. สถานีตำรวจทางหลวง (ส.ทล.)
    # 4. หน่วยบริการ / ตู้สายตรวจ
    # 5. ตำแหน่ง / สิทธิ์ในระบบ
    # 6. ต้นทางที่ส่งมาช่วยราชการ
    # 7. ปลายทางที่ไปช่วยราชการ
    # 8. วันที่เริ่มช่วยราชการ
    # 9. วันที่สิ้นสุดช่วยราชการ
    headers = [
        "ลำดับ",
        "ยศ - ชื่อ - นามสกุล",
        "สถานีตำรวจทางหลวง (ส.ทล.)",
        "หน่วยบริการ / ตู้สายตรวจ",
        "ตำแหน่ง / สิทธิ์ในระบบ",
        "ต้นทางที่ส่งมาช่วยราชการ",
        "ปลายทางที่ไปช่วยราชการ",
        "วันที่เริ่มช่วยราชการ",
        "วันที่สิ้นสุดช่วยราชการ"
    ]

    for st in stations:
        ws = wb.create_sheet(title=st["sheet"])
        ws.views.sheetView[0].showGridLines = True

        # Title Block
        ws.merge_cells("A1:I1")
        ws["A1"] = f"แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่ - {st['fullName']}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")

        ws.merge_cells("A2:I2")
        ws["A2"] = "* กรุณากรอก ยศ-ชื่อ-สกุล หน่วยบริการ ตำแหน่ง (กรณีช่วยราชการให้ระบุหน่วยงานต้นทาง/ปลายทาง และวันเริ่ม-สิ้นสุด)"
        ws["A2"].font = note_font
        ws["A2"].alignment = Alignment(vertical="center")

        # Header Row
        ws.row_dimensions[4].height = 28
        for col_num, h_title in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = h_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Sample Row (Row 5)
        sample_row = [
            1,
            "ด.ต. สมชาย สายตรวจ (ตัวอย่าง)",
            st["fullName"],
            st["units"][0] if st["units"] else "หน่วยบริการดอนแก้ว",
            "ผู้ปฏิบัติประจำหน่วย",
            "-",
            "-",
            "-",
            "-"
        ]
        ws.row_dimensions[5].height = 24
        for col_num, val in enumerate(sample_row, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = val
            cell.font = sample_font
            cell.border = thin_border
            if col_num in [1, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Blank rows for user filling (Row 6 to 50)
        for seq in range(2, 51):
            r = seq + 4
            ws.row_dimensions[r].height = 22

            row_values = [
                seq,
                "", # FullName
                st["fullName"],
                "", # Unit
                "ผู้ปฏิบัติประจำหน่วย" if seq > 2 else ("สิบเวร / Admin สถานี" if seq == 2 else "หัวหน้าหน่วยบริการ"),
                "", # ต้นทางที่ส่งมาช่วยราชการ
                "", # ปลายทางที่ไปช่วยราชการ
                "", # วันที่เริ่ม
                ""  # วันที่สิ้นสุด
            ]

            for c_num, val in enumerate(row_values, 1):
                cell = ws.cell(row=r, column=c_num)
                cell.value = val
                cell.font = body_font
                cell.border = thin_border
                
                if c_num in [1, 8, 9]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Unit Data Validation
        if st["units"]:
            units_str = ",".join(st["units"])
            dv_unit = DataValidation(type="list", formula1=f'"{units_str}"', allow_blank=True)
            dv_unit.error = "กรุณาเลือก หน่วยบริการ จากรายการที่มีให้"
            dv_unit.errorTitle = "ข้อมูลไม่ถูกต้อง"
            ws.add_data_validation(dv_unit)
            dv_unit.add("D5:D54")

        # Role Data Validation
        dv_role = DataValidation(type="list", formula1='"ผู้ปฏิบัติประจำหน่วย,สิบเวร / Admin สถานี,หัวหน้าหน่วยบริการ,ฝอ.กก.,ผกก."', allow_blank=True)
        ws.add_data_validation(dv_role)
        dv_role.add("E5:E54")

        # Column widths
        col_widths = {
            "A": 8,
            "B": 28,
            "C": 34,
            "D": 24,
            "E": 24,
            "F": 25,
            "G": 25,
            "H": 20,
            "I": 20
        }
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w

    # Add Guide Sheet at end
    ws_guide = wb.create_sheet(title="คำแนะนำการกรอกข้อมูล")
    ws_guide.views.sheetView[0].showGridLines = True
    ws_guide["A1"] = f"คำแนะนำการกรอกแบบฟอร์มข้อมูลเจ้าหน้าที่ - กก.{div_num} บก.ทล."
    ws_guide["A1"].font = title_font

    guides = [
        ("ชื่อคอลัมน์", "ความสำคัญ", "รายละเอียดและข้อแนะนำ"),
        ("ลำดับ", "อัตโนมัติ", "ลำดับที่ 1 - 50 ประจำสถานี"),
        ("ยศ - ชื่อ - นามสกุล", "บังคับ", "ระบุยศ ชื่อ และนามสกุลให้ครบถ้วน เช่น ด.ต. สมชาย สายตรวจ"),
        ("สถานีตำรวจทางหลวง (ส.ทล.)", "อัตโนมัติ", "ระบบเติมชื่อสถานีตามชื่อแท็บชีตให้อัตโนมัติ"),
        ("หน่วยบริการ / ตู้สายตรวจ", "บังคับ", "กดเลือกหน่วยบริการประจำสถานีนั้นๆ จากเมนู Dropdown"),
        ("ตำแหน่ง / สิทธิ์ในระบบ", "บังคับ", "กดเลือกสิทธิ์ในระบบ HWPD Next Gen:\n - ผู้ปฏิบัติประจำหน่วย: สายตรวจ/ผู้บันทึกรายงาน\n - สิบเวร / Admin สถานี: อนุมัติรายงานประจำสถานี\n - หัวหน้าหน่วยบริการ: หัวหน้าตู้/หน่วยบริการ\n - ฝอ.กก.: ฝ่ายอำนวยการระดับ กก.\n - ผกก.: ผู้กำกับการ"),
        ("ต้นทางที่ส่งมาช่วยราชการ", "กรณีมาช่วยราชการ", "ระบุชื่อหน่วยงานเดิมต้นทาง กรณีข้าราชการตำรวจมาช่วยราชการที่หน่วยนี้ (เช่น ส.ทล.2 กก.5)"),
        ("ปลายทางที่ไปช่วยราชการ", "กรณีไปช่วยราชการ", "ระบุชื่อหน่วยงานปลายทาง กรณีข้าราชการตำรวจของหน่วยนี้ไปช่วยราชการที่อื่น (เช่น บก.ปปป.)"),
        ("วันที่เริ่ม / สิ้นสุดช่วยราชการ", "กรณีช่วยราชการ", "ระบุวันที่ในรูปแบบ YYYY-MM-DD (เช่น 2026-08-01) หากไม่ได้ช่วยราชการให้ละเว้นไว้")
    ]

    ws_guide.row_dimensions[3].height = 26
    for c, h in enumerate(guides[0], 1):
        cell = ws_guide.cell(row=3, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row_data in enumerate(guides[1:], 4):
        ws_guide.row_dimensions[r_idx].height = 35
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_guide.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if c_idx == 1:
                cell.font = Font(name="Tahoma", size=10, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_guide.column_dimensions["A"].width = 30
    ws_guide.column_dimensions["B"].width = 18
    ws_guide.column_dimensions["C"].width = 75

    # Remove initial blank sheet
    wb.remove(default_sheet)

    file_name = f"แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่_กก{div_num}_บก.ทล.xlsx"
    file_path = os.path.join(output_dir, file_name)
    wb.save(file_path)
    print(f"Generated EXACT 9-Column Workbook: {file_name}")


if __name__ == "__main__":
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    out_dir = os.path.join(base_dir, "แบบฟอร์มแยกตาม_กก")
    os.makedirs(out_dir, exist_ok=True)

    for div_num, stations in STATIONS_DATA.items():
        build_exact_division_workbook(div_num, stations, out_dir)

    print(f"\nSUCCESS! ALL 8 WORKBOOKS GENERATED WITH EXACT 9 COLUMNS IN: {out_dir}")
