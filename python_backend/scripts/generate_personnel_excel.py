# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "แบบฟอร์มกรอกข้อมูลเจ้าหน้าที่"
ws.views.sheetView[0].showGridLines = True

header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
header_font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
sample_font = Font(name="Tahoma", size=10, italic=True, color="555555")
body_font = Font(name="Tahoma", size=10)
title_font = Font(name="Tahoma", size=14, bold=True, color="1E3A8A")
note_font = Font(name="Tahoma", size=10, italic=True, color="DC2626")

thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)

headers = [
    "ลำดับ",
    "ยศ - ชื่อ - นามสกุล",
    "สถานีตำรวจทางหลวง (ส.ทล.)",
    "หน่วยบริการ / ตู้สายตรวจ",
    "ตำแหน่ง / สิทธิ์ในระบบ",
    "ต้นทางที่ส่งมาช่วยราชการ",
    "ปลายทางที่ไปช่วยราชการ",
    "วันที่เริ่มช่วยราชการ",
    "วันที่สิ้นสุดช่วยราชการ"
]

ws.merge_cells("A1:I1")
ws["A1"] = "แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่ตำรวจทางหลวง (สำหรับระบบ HWPD Next Gen)"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(vertical="center")

ws.merge_cells("A2:I2")
ws["A2"] = "* กรุณากรอก ยศ-ชื่อ-สกุล สถานี สังกัดหน่วยบริการ ตำแหน่ง (กรณีช่วยราชการให้ระบุหน่วยงานต้นทาง/ปลายทาง และวันเริ่ม-สิ้นสุด)"
ws["A2"].font = note_font
ws["A2"].alignment = Alignment(vertical="center")

ws.row_dimensions[4].height = 28
for col_num, h_title in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.value = h_title
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

sample_row = [
    1,
    "ด.ต. สมชาย สายตรวจ (ตัวอย่าง)",
    "ส.ทล.1 กก.5 บก.ทล. (เชียงใหม่)",
    "หน่วยบริการดอนแก้ว",
    "ผู้ปฏิบัติประจำหน่วย",
    "-",
    "-",
    "-",
    "-"
]
ws.row_dimensions[5].height = 24
for col_num, val in enumerate(sample_row, 1):
    cell = ws.cell(row=5, column=col_num)
    cell.value = val
    cell.font = sample_font
    cell.border = thin_border
    if col_num in [1, 8, 9]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")

for seq in range(2, 101):
    r = seq + 4
    ws.row_dimensions[r].height = 22
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.font = body_font
        cell.border = thin_border
        if c == 1:
            cell.value = seq
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c in [8, 9]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

dv_role = DataValidation(type="list", formula1='"ผู้ปฏิบัติประจำหน่วย,สิบเวร / Admin สถานี,หัวหน้าหน่วยบริการ,ฝอ.กก.,ผกก."', allow_blank=True)
ws.add_data_validation(dv_role)
dv_role.add("E5:E104")

col_widths = {"A": 8, "B": 28, "C": 34, "D": 24, "E": 24, "F": 25, "G": 25, "H": 20, "I": 20}
for col_letter, w in col_widths.items():
    ws.column_dimensions[col_letter].width = w

base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
file_path = os.path.join(base_dir, "แบบฟอร์มรวบรวมข้อมูลเจ้าหน้าที่_บก.ทล.xlsx")
wb.save(file_path)
print(f"Updated single master workbook: {file_path}")
